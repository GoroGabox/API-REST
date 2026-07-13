import io
import json
import os
import re
import tempfile

from rest_framework import viewsets, status, permissions as drf_permissions
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from django.conf import settings
from django.db import transaction as db_transaction
from django.db.models import Prefetch
from django.http import StreamingHttpResponse, HttpResponse
from django_filters.rest_framework import DjangoFilterBackend

from accounts.permissions import IsAdmin, ReadOnlyOrAdmin, PublicReadOrAdmin, is_admin, is_director, is_estudiante
from .models import Escuela, Curso, Leccion, Ejercicio, Glosario, Categoria, Unidad, Recurso
from .serializers import (
    EscuelaSerializer,
    CursoSerializer,
    LeccionSerializer,
    LeccionDetalleSerializer,
    EjercicioSerializer,
    EjercicioConRespuestaSerializer,
    GlosarioSerializer,
    CategoriaSerializer,
    UnidadSerializer,
    RecursoSerializer,
)


LECCION_ORDERING = ('curso_id', 'unidad__orden', 'unidad_id', 'posicion', 'id')


# Campos que un director puede modificar en su propia Escuela via PATCH.
# Todo lo demás (basic_key, professional_key, basic_access, professional_access)
# queda reservado a admin y procesos internos (pagos).
DIRECTOR_EDITABLE_ESCUELA_FIELDS = {"nombre", "email", "telefono", "direccion"}


# ---------------------------------------------------------------------------
# Carga masiva de ejercicios (EjercicioViewSet.bulk_upload / bulk_template)
# ---------------------------------------------------------------------------
# Orden canónico de columnas de la plantilla .xlsx.
BULK_COLUMNS = [
    "id", "pregunta",
    "opcion_a", "opcion_b", "opcion_c", "opcion_d", "opcion_e", "opcion_f",
    "respuesta", "categoria_id", "curso_id", "leccion_id", "imagen",
]
BULK_REQUIRED_COLUMNS = ["pregunta", "opcion_a", "opcion_b", "respuesta", "categoria_id"]
# Alias aceptados en la cabecera → nombre canónico.
BULK_ALIASES = {
    "categoria": "categoria_id", "categoria_id": "categoria_id",
    "curso": "curso_id", "curso_id": "curso_id",
    "leccion": "leccion_id", "leccion_id": "leccion_id",
    "pregunta_texto": "pregunta", "imagen_url": "imagen",
}
BULK_UPLOAD_MAX_BYTES = 10 * 1024 * 1024  # 10 MB


def _norm_header(name):
    """Normaliza un nombre de columna: minúsculas, sin espacios/acentos raros."""
    if name is None:
        return ""
    return str(name).strip().lower().replace(" ", "_")


def _clean(value):
    """Celda → str limpio o None. Trata "", "NULL", "-", "N/A" como vacío."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.upper() in ("NULL", "NONE", "N/A", "-", "—"):
        return None
    return s


def _as_int(value):
    """Celda → int o None (tolera floats de Excel como 13.0 y strings)."""
    s = _clean(value)
    if s is None:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


class _EscuelaScopedPermission(drf_permissions.BasePermission):
    """Admin: full CRUD. Director: GET + PATCH sobre su propia escuela.
    Estudiante: sin acceso.
    """
    def has_permission(self, request, view):
        if not (request.user and request.user.is_authenticated):
            return False
        if is_admin(request.user):
            return True
        if is_director(request.user):
            # POST/DELETE prohibidos para director.
            return request.method in ("GET", "HEAD", "OPTIONS", "PATCH", "PUT")
        return False

    def has_object_permission(self, request, view, obj):
        if is_admin(request.user):
            return True
        if is_director(request.user):
            return obj.pk == request.user.escuela_id
        return False


class EscuelaViewSet(viewsets.ModelViewSet):
    """Admin: full CRUD sobre todas las escuelas.

    Director: puede VER y EDITAR (campos limitados) su propia escuela.
    Estudiante: no accede aquí.
    """
    queryset = Escuela.objects.all()
    serializer_class = EscuelaSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['email', 'basic_access', 'professional_access']
    permission_classes = [_EscuelaScopedPermission]

    def get_queryset(self):
        user = self.request.user
        if is_admin(user):
            return Escuela.objects.all()
        if is_director(user) and user.escuela_id:
            return Escuela.objects.filter(pk=user.escuela_id)
        return Escuela.objects.none()

    def update(self, request, *args, **kwargs):
        return self._perform_scoped_write(request, partial=kwargs.pop("partial", False), *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self._perform_scoped_write(request, partial=True, *args, **kwargs)

    def _perform_scoped_write(self, request, partial=False, *args, **kwargs):
        """PATCH/PUT: si es director, restringe los campos editables.
        Admin puede editar cualquier campo.
        """
        instance = self.get_object()  # 404/403 aplicados por has_object_permission
        data = request.data
        if is_director(request.user) and not is_admin(request.user):
            data = {k: v for k, v in dict(data).items() if k in DIRECTOR_EDITABLE_ESCUELA_FIELDS}
            if not data:
                return Response(
                    {"detail": f"Campos editables por director: {sorted(DIRECTOR_EDITABLE_ESCUELA_FIELDS)}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        serializer = self.get_serializer(instance, data=data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class ProgresoEstudianteDetalleView(APIView):
    """GET /api/v1/schools/<school_id>/estudiantes/<user_id>/progreso/

    Devuelve el progreso detallado del estudiante en todos los cursos donde
    tiene acceso: lecciones completadas (existencia de EstudianteLeccion),
    pruebas realizadas, certificado si existe.

    Permisos:
    - admin: cualquier estudiante.
    - director: solo estudiantes de su escuela (school_id debe ser la suya).
    - estudiante: solo sobre sí mismo.
    """
    permission_classes = [drf_permissions.IsAuthenticated]

    def get(self, request, school_id, user_id):
        from accounts.models import Usuario, EstudianteLeccion, Prueba, Certificado
        from sales.models import EstudianteCurso

        try:
            school_id = int(school_id)
            user_id = int(user_id)
        except (TypeError, ValueError):
            return Response({"detail": "IDs inválidos."}, status=status.HTTP_400_BAD_REQUEST)

        if not is_admin(request.user):
            if is_director(request.user):
                if request.user.escuela_id != school_id:
                    return Response({"detail": "No autorizado para esa escuela."}, status=status.HTTP_403_FORBIDDEN)
            elif is_estudiante(request.user):
                if request.user.id != user_id:
                    return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)
            else:
                return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)

        try:
            estudiante = Usuario.objects.get(pk=user_id)
        except Usuario.DoesNotExist:
            return Response({"detail": "Estudiante no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        if estudiante.escuela_id != school_id and not is_admin(request.user):
            return Response({"detail": "El estudiante no pertenece a esta escuela."}, status=status.HTTP_404_NOT_FOUND)

        inscripciones = (
            EstudianteCurso.objects
            .filter(estudiante_id=estudiante)
            .select_related("curso_id", "access_key_id")
        )
        curso_ids = [ec.curso_id_id for ec in inscripciones]

        # Batch queries por curso para minimizar N+1.
        lecciones_por_curso = {}
        for l in Leccion.objects.filter(curso_id__in=curso_ids):
            lecciones_por_curso.setdefault(l.curso_id, []).append(l)

        completadas = set(
            EstudianteLeccion.objects
            .filter(estudiante=estudiante, curso_id__in=curso_ids)
            .values_list("curso_id", "leccion_id", "updated_at")
        )
        fecha_leccion = {(c, l): t for (c, l, t) in completadas}
        set_completadas = set((c, l) for (c, l, _) in completadas)

        pruebas_por_curso = {}
        for p in Prueba.objects.filter(estudiante=estudiante, curso_id__in=curso_ids):
            pruebas_por_curso.setdefault(p.curso_id, []).append(p)

        certs_por_curso = {
            c.curso_id: c
            for c in Certificado.objects.filter(estudiante=estudiante, curso_id__in=curso_ids)
        }

        cursos_data = []
        for ec in inscripciones:
            curso = ec.curso_id
            key = ec.access_key_id
            lecs = lecciones_por_curso.get(curso.id, [])
            lecciones_detalle = [
                {
                    "id": l.id,
                    "nombre": l.nombre,
                    "unidad": l.unidad_id,
                    "posicion": l.posicion,
                    "completada": (curso.id, l.id) in set_completadas,
                    "fecha_completada": fecha_leccion.get((curso.id, l.id)),
                }
                for l in lecs
            ]
            lecciones_completadas = sum(1 for x in lecciones_detalle if x["completada"])
            pruebas = [
                {
                    "id": p.id,
                    "tipo": p.tipo,
                    "fecha": p.fecha,
                    "score": float(p.score) if p.score is not None else None,
                    "aprobado": p.aprobado,
                }
                for p in pruebas_por_curso.get(curso.id, [])
            ]
            cert = certs_por_curso.get(curso.id)
            certificado = None
            if cert is not None:
                certificado = {
                    "id": cert.id,
                    "codigo": str(cert.codigo),
                    "fecha_emision": cert.fecha_emision,
                }
            cursos_data.append({
                "curso_id": curso.id,
                "curso_nombre": curso.nombre,
                "acceso": {
                    "valid_from": key.valid_from if key else None,
                    "valid_until": key.valid_until if key else None,
                    "status": key.status if key else None,
                },
                "lecciones_total": len(lecs),
                "lecciones_completadas": lecciones_completadas,
                "lecciones_detalle": lecciones_detalle,
                "pruebas": pruebas,
                "certificado": certificado,
            })

        return Response({
            "estudiante": {
                "id": estudiante.id,
                "nombre": estudiante.nombre,
                "apellido": estudiante.apellido,
                "email": estudiante.email,
            },
            "cursos": cursos_data,
        })


class CertificadosPorEscuelaView(APIView):
    """GET /api/v1/schools/<school_id>/certificados/

    Lista los certificados emitidos a estudiantes de la escuela.
    Filtros: ?curso=<id>, ?desde=YYYY-MM-DD, ?hasta=YYYY-MM-DD

    Permisos: admin (cualquier escuela) o director (solo la suya).
    """
    permission_classes = [drf_permissions.IsAuthenticated]

    def get(self, request, school_id):
        from accounts.models import Certificado

        try:
            school_id = int(school_id)
        except (TypeError, ValueError):
            return Response({"detail": "school_id inválido."}, status=status.HTTP_400_BAD_REQUEST)

        if not is_admin(request.user):
            if not is_director(request.user) or request.user.escuela_id != school_id:
                return Response({"detail": "No autorizado."}, status=status.HTTP_403_FORBIDDEN)

        qs = (
            Certificado.objects
            .filter(estudiante__escuela_id=school_id)
            .select_related("estudiante", "curso", "prueba")
        )

        curso_param = request.query_params.get("curso")
        if curso_param:
            try:
                qs = qs.filter(curso_id=int(curso_param))
            except ValueError:
                return Response({"detail": "curso inválido."}, status=status.HTTP_400_BAD_REQUEST)

        desde = request.query_params.get("desde")
        hasta = request.query_params.get("hasta")
        if desde:
            qs = qs.filter(fecha_emision__gte=desde)
        if hasta:
            qs = qs.filter(fecha_emision__lte=hasta)

        qs = qs.order_by("-fecha_emision", "-id")
        results = [
            {
                "id": c.id,
                "codigo": str(c.codigo),
                "estudiante": {
                    "id": c.estudiante_id,
                    "nombre": f"{c.estudiante.nombre} {c.estudiante.apellido}".strip(),
                    "email": c.estudiante.email,
                },
                "curso": {
                    "id": c.curso_id,
                    "nombre": c.curso.nombre if c.curso else None,
                },
                "fecha_emision": c.fecha_emision,
                "prueba_id": c.prueba_id,
                "score": float(c.prueba.score) if c.prueba and c.prueba.score is not None else None,
            }
            for c in qs
        ]

        return Response({"count": len(results), "results": results})


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _enviar_credenciales_estudiante(email, nombre, password):
    """Envía credenciales temporales al estudiante recién creado (best-effort)."""
    try:
        from django.core.mail import send_mail
        send_mail(
            subject="Tu cuenta AutoTest",
            message=(
                f"Hola {nombre or 'estudiante'},\n\n"
                f"Se creó tu cuenta en AutoTest.\n"
                f"Correo: {email}\n"
                f"Contraseña temporal: {password}\n\n"
                f"Te recomendamos cambiarla en tu primer inicio de sesión."
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=True,
        )
    except Exception:
        pass


def vincular_o_crear_estudiante(email, nombre, apellido, escuela_id):
    """Vincula (o crea) un estudiante en `escuela_id`. Maneja su propia
    transacción y, al crear, envía credenciales por email.

    Devuelve dict {status, email, user_id?, detail?} con
    status ∈ {created, linked, already_linked, error}.
    """
    from accounts.models import Usuario

    email = (email or "").strip().lower()
    if not email:
        return {"status": "error", "email": email, "detail": "email vacío."}
    if not _EMAIL_RE.match(email):
        return {"status": "error", "email": email, "detail": "email inválido."}

    nombre = (nombre or "").strip()
    apellido = (apellido or "").strip()

    with db_transaction.atomic():
        existing = Usuario.objects.filter(email__iexact=email).select_for_update().first()

        if existing is not None:
            if existing.is_staff or existing.is_superuser or existing.is_director:
                return {"status": "error", "email": email, "detail": "usuario administrativo."}
            if existing.escuela_id == escuela_id:
                return {"status": "already_linked", "email": existing.email, "user_id": existing.id}
            existing.escuela_id = escuela_id
            existing.is_estudiante = True
            existing.is_active = True
            existing.save(update_fields=["escuela", "is_estudiante", "is_active"])
            return {"status": "linked", "email": existing.email, "user_id": existing.id}

        import secrets
        random_password = secrets.token_urlsafe(12)
        new_user = Usuario.objects.create_user(
            email=email,
            nombre=nombre or "Estudiante",
            apellido=apellido or "",
            password=random_password,
            is_estudiante=True,
        )
        new_user.is_active = True
        new_user.escuela_id = escuela_id
        new_user.save(update_fields=["is_active", "escuela"])
        _enviar_credenciales_estudiante(email, nombre, random_password)
        return {"status": "created", "email": new_user.email, "user_id": new_user.id}


def _resolver_escuela_objetivo(request):
    """Escuela destino para vincular/importar. Devuelve (escuela_id, error|None).

    director → su propia escuela; admin → `escuela_id` del body.
    """
    if is_director(request.user):
        if not request.user.escuela_id:
            return None, Response(
                {"detail": "El director no tiene escuela asignada."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        return request.user.escuela_id, None
    escuela_id = request.data.get("escuela_id") or request.data.get("escuela")
    if not escuela_id:
        return None, Response(
            {"detail": "escuela_id requerido cuando el caller es admin."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if not Escuela.objects.filter(pk=escuela_id).exists():
        return None, Response({"detail": "Escuela no existe."}, status=status.HTTP_404_NOT_FOUND)
    return escuela_id, None


class VincularEstudianteView(APIView):
    """POST /api/v1/schools/vincular-estudiante/  {email, nombre?, apellido?}

    Vincula (o crea) un estudiante a la escuela del director autenticado.

    Respuestas:
      201 { status: "created",        user_id, email }
      200 { status: "linked",         user_id, email }
      200 { status: "already_linked", user_id, email }
      403 { detail: ... }             email de admin/director/staff
      400 { detail: "email requerido." }
    """
    permission_classes = [drf_permissions.IsAuthenticated]

    def post(self, request):
        if not (is_director(request.user) or is_admin(request.user)):
            return Response({"detail": "Solo director o admin."}, status=status.HTTP_403_FORBIDDEN)

        escuela_id, err = _resolver_escuela_objetivo(request)
        if err is not None:
            return err

        if not (request.data.get("email") or "").strip():
            return Response({"detail": "email requerido."}, status=status.HTTP_400_BAD_REQUEST)

        result = vincular_o_crear_estudiante(
            request.data.get("email"), request.data.get("nombre"),
            request.data.get("apellido"), escuela_id,
        )
        if result["status"] == "error":
            if "administrativo" in (result.get("detail") or ""):
                return Response(
                    {"detail": "No se puede vincular usuario administrativo."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            return Response(
                {"detail": result.get("detail") or "email inválido."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        code = status.HTTP_201_CREATED if result["status"] == "created" else status.HTTP_200_OK
        return Response(result, status=code)


class ImportarEstudiantesView(APIView):
    """POST /api/v1/schools/importar-estudiantes/
       { estudiantes: [{email, nombre?, apellido?}], escuela_id? }

    Vincula/crea estudiantes en lote (carga masiva). Cada fila se procesa de
    forma independiente; una fila con error no aborta el resto. Devuelve un
    resumen agregado + el detalle por fila.
    """
    permission_classes = [drf_permissions.IsAuthenticated]
    MAX_ROWS = 300

    def post(self, request):
        if not (is_director(request.user) or is_admin(request.user)):
            return Response({"detail": "Solo director o admin."}, status=status.HTTP_403_FORBIDDEN)

        escuela_id, err = _resolver_escuela_objetivo(request)
        if err is not None:
            return err

        rows = request.data.get("estudiantes")
        if not isinstance(rows, list) or not rows:
            return Response(
                {"detail": "Envía 'estudiantes' como una lista no vacía."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(rows) > self.MAX_ROWS:
            return Response(
                {"detail": f"Máximo {self.MAX_ROWS} estudiantes por importación."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        results = []
        summary = {"created": 0, "linked": 0, "already_linked": 0, "error": 0}
        seen = set()
        for row in rows:
            if not isinstance(row, dict):
                results.append({"status": "error", "email": "", "detail": "fila inválida."})
                summary["error"] += 1
                continue
            email_norm = (row.get("email") or "").strip().lower()
            if email_norm and email_norm in seen:
                results.append({"status": "error", "email": email_norm, "detail": "duplicado en el archivo."})
                summary["error"] += 1
                continue
            if email_norm:
                seen.add(email_norm)
            res = vincular_o_crear_estudiante(
                row.get("email"), row.get("nombre"), row.get("apellido"), escuela_id,
            )
            summary[res["status"]] = summary.get(res["status"], 0) + 1
            results.append(res)

        return Response({"summary": summary, "results": results}, status=status.HTTP_200_OK)


class DesvincularEstudianteView(APIView):
    """POST /api/v1/schools/desvincular-estudiante/ {user_id}

    Quita a un estudiante de la escuela del director: pone `escuela=None` y
    retira su acceso a los cursos otorgados por la escuela — marca las
    AccessKey como 'revoked', libera cupos de suscripción (decrementa
    *_seats_used) y borra los EstudianteCurso.

    Permisos:
    - admin: cualquier estudiante.
    - director: solo estudiantes de su escuela.
    Rechaza usuarios administrativos.
    """
    permission_classes = [drf_permissions.IsAuthenticated]

    def post(self, request):
        if not (is_director(request.user) or is_admin(request.user)):
            return Response({"detail": "Solo director o admin."}, status=status.HTTP_403_FORBIDDEN)

        user_id = request.data.get("user_id")
        if not user_id:
            return Response({"detail": "user_id requerido."}, status=status.HTTP_400_BAD_REQUEST)

        from accounts.models import Usuario
        from sales.models import EstudianteCurso

        with db_transaction.atomic():
            try:
                target = Usuario.objects.select_for_update().get(pk=user_id)
            except Usuario.DoesNotExist:
                return Response({"detail": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

            if target.is_staff or target.is_superuser or target.is_director:
                return Response(
                    {"detail": "No se puede desvincular un usuario administrativo."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            if is_director(request.user) and target.escuela_id != request.user.escuela_id:
                return Response(
                    {"detail": "El estudiante no pertenece a tu escuela."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            escuela = (
                Escuela.objects.select_for_update().filter(pk=target.escuela_id).first()
                if target.escuela_id else None
            )

            inscripciones = (
                EstudianteCurso.objects
                .select_related("curso_id", "access_key_id")
                .filter(estudiante_id=target)
            )
            for ec in inscripciones:
                ak = ec.access_key_id
                # Compra individual: el acceso es del estudiante (lo pagó), NO
                # de la escuela. Al salir de la escuela lo conserva intacto.
                if ak and ak.origen == "purchase":
                    continue
                if ak and ak.origen == "seat" and escuela is not None:
                    if ec.curso_id.is_profesional:
                        if escuela.professional_seats_used > 0:
                            escuela.professional_seats_used -= 1
                    else:
                        if escuela.basic_seats_used > 0:
                            escuela.basic_seats_used -= 1
                if ak:
                    ak.status = "revoked"
                    ak.save(update_fields=["status"])
                ec.delete()

            if escuela is not None:
                escuela.save()

            target.escuela = None
            target.save(update_fields=["escuela"])

        return Response({"status": "unlinked", "user_id": target.id}, status=status.HTTP_200_OK)


class EditarEstudianteView(APIView):
    """PATCH /api/v1/schools/estudiantes/<user_id>/
       { nombre?, apellido?, telefono?, rut?, direccion?, email? }

    Edita datos del roster de un estudiante. NO permite tocar roles ni estado.
    Director: solo estudiantes de su escuela y no administrativos. El email
    (identidad de login) se valida en formato y unicidad.
    """
    permission_classes = [drf_permissions.IsAuthenticated]
    ALLOWED = ("nombre", "apellido", "telefono", "rut", "direccion")

    def patch(self, request, user_id):
        if not (is_director(request.user) or is_admin(request.user)):
            return Response({"detail": "Solo director o admin."}, status=status.HTTP_403_FORBIDDEN)

        from accounts.models import Usuario
        try:
            target = Usuario.objects.get(pk=user_id)
        except Usuario.DoesNotExist:
            return Response({"detail": "Usuario no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        if target.is_staff or target.is_superuser or target.is_director:
            return Response(
                {"detail": "No se puede editar un usuario administrativo."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if is_director(request.user) and target.escuela_id != request.user.escuela_id:
            return Response(
                {"detail": "El estudiante no pertenece a tu escuela."},
                status=status.HTTP_403_FORBIDDEN,
            )

        update_fields = []
        # nombre/apellido no pueden quedar vacíos si vienen en el payload.
        for f in self.ALLOWED:
            if f in request.data:
                val = (request.data.get(f) or "").strip()
                if f in ("nombre", "apellido") and not val:
                    return Response({"detail": f"El campo {f} no puede quedar vacío."},
                                    status=status.HTTP_400_BAD_REQUEST)
                setattr(target, f, val)
                update_fields.append(f)

        # Email: identidad de login → validar formato + unicidad.
        if "email" in request.data:
            new_email = (request.data.get("email") or "").strip().lower()
            if not new_email or not _EMAIL_RE.match(new_email):
                return Response({"detail": "Email inválido."}, status=status.HTTP_400_BAD_REQUEST)
            if Usuario.objects.filter(email__iexact=new_email).exclude(pk=target.pk).exists():
                return Response({"detail": "Ya existe un usuario con ese email."},
                                status=status.HTTP_409_CONFLICT)
            if new_email != (target.email or "").lower():
                target.email = new_email
                update_fields.append("email")

        if not update_fields:
            return Response({"detail": "No se enviaron cambios."}, status=status.HTTP_400_BAD_REQUEST)

        target.save(update_fields=update_fields)
        return Response(
            {
                "id": target.id,
                "nombre": target.nombre,
                "apellido": target.apellido,
                "email": target.email,
                "telefono": target.telefono,
                "rut": target.rut,
                "direccion": target.direccion,
            },
            status=status.HTTP_200_OK,
        )


class CursoViewSet(viewsets.ModelViewSet):
    queryset = Curso.objects.all()
    serializer_class = CursoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['costo']
    permission_classes = [PublicReadOrAdmin]

    @action(detail=True, methods=['get'], url_path='units')
    def units(self, request, pk=None):
        """GET /api/v1/schools/courses/<id>/units/ — unidades del curso con sus lecciones."""
        curso = self.get_object()
        lecciones_ordenadas = Leccion.objects.select_related('categoria').order_by('posicion', 'id')
        unidades = (
            Unidad.objects.filter(curso=curso)
            .prefetch_related(Prefetch('lecciones', queryset=lecciones_ordenadas))
            .order_by('orden', 'id')
        )
        return Response(UnidadSerializer(unidades, many=True).data)


class LeccionViewSet(viewsets.ModelViewSet):
    queryset = Leccion.objects.all()
    serializer_class = LeccionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['curso', 'unidad', 'posicion', 'tipo']
    permission_classes = [ReadOnlyOrAdmin]
    pagination_class = None

    def get_queryset(self):
        return (
            Leccion.objects
            .select_related('curso', 'unidad', 'categoria')
            .order_by(*LECCION_ORDERING)
        )

    def get_serializer_class(self):
        # Detalle (con contenido + transcripción) en:
        # - retrieve
        # - list?curso=<id> (el frontend del curso consume la lista completa)
        # - escrituras (create/update/partial_update): el admin edita el cuerpo
        #   generado por el pipeline para producir la lección final. Sin esto,
        #   un PATCH con `contenido`/`transcripcion` se descartaría en silencio.
        write_actions = ('create', 'update', 'partial_update')
        if (
            self.action in ('retrieve', *write_actions)
            or (self.action == 'list' and self.request.query_params.get('curso'))
        ):
            return LeccionDetalleSerializer
        return LeccionSerializer


class UnidadViewSet(viewsets.ModelViewSet):
    queryset = Unidad.objects.all()
    serializer_class = UnidadSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['curso']
    permission_classes = [ReadOnlyOrAdmin]


class EjercicioViewSet(viewsets.ModelViewSet):
    queryset = Ejercicio.objects.all()
    serializer_class = EjercicioSerializer
    permission_classes = [ReadOnlyOrAdmin]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['categoria', 'curso', 'leccion']

    def get_serializer_class(self):
        """Admin ve y escribe `respuesta`; el resto recibe el serializer
        público que oculta la respuesta correcta."""
        if is_admin(self.request.user):
            return EjercicioConRespuestaSerializer
        return EjercicioSerializer

    # ------------------------------------------------------------------
    # Carga masiva desde .xlsx
    # ------------------------------------------------------------------
    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-template",
        permission_classes=[IsAdmin],
    )
    def bulk_template(self, request):
        """GET /api/v1/schools/exercices/bulk-template/

        Devuelve un .xlsx con la cabecera exacta que espera `bulk_upload`
        más una fila de ejemplo, para que el admin no tenga que adivinar el
        formato. Descarga como adjunto.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ejercicios"
        ws.append(BULK_COLUMNS)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F46E5")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.append([
            1,
            "Su vehículo se desvía hacia un lado cuando usted frena. Usted debería:",
            "Cambiar los neumáticos de un lado hacia el otro y viceversa.",
            "Bombear el pedal al frenar.",
            "Usar su freno de mano.",
            "Consultar con su mecánico lo antes posible.",
            "NULL", "NULL", "D", 13, 1, "NULL", "http://placeholder.url",
        ])
        # Anchos legibles.
        widths = [5, 60, 30, 30, 30, 30, 12, 12, 10, 12, 10, 12, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        resp["Content-Disposition"] = (
            'attachment; filename="plantilla_ejercicios.xlsx"'
        )
        return resp

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-upload",
        permission_classes=[IsAdmin],
        parser_classes=[MultiPartParser, FormParser],
    )
    def bulk_upload(self, request):
        """POST /api/v1/schools/exercices/bulk-upload/

        multipart/form-data:
          - file: .xlsx con las columnas de `bulk_template`.
          - dry_run: "true"|"1" → valida y reporta sin escribir en BD.

        `respuesta` acepta la LETRA de la opción correcta (A-F) o el texto
        exacto de una opción; se persiste como el TEXTO (coherente con el
        CRUD y con la corrección del quiz en la app).

        Responde 200 con un resumen por fila; las filas inválidas se omiten
        (no abortan el lote) salvo que se pida `dry_run`.
        """
        import openpyxl

        upload = request.FILES.get("file") or request.FILES.get("archivo")
        if not upload:
            return Response(
                {"detail": "Adjunta el archivo .xlsx en el campo 'file'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not str(upload.name).lower().endswith((".xlsx", ".xlsm")):
            return Response(
                {"detail": "El archivo debe ser .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if upload.size > BULK_UPLOAD_MAX_BYTES:
            return Response(
                {"detail": f"El archivo supera {BULK_UPLOAD_MAX_BYTES // (1024 * 1024)} MB."},
                status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            )
        dry_run = str(request.data.get("dry_run", "")).lower() in ("true", "1", "on", "yes")
        skip_duplicates = str(request.data.get("skip_duplicates", "")).lower() in ("true", "1", "on", "yes")

        try:
            wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
        except Exception:
            return Response(
                {"detail": "No pudimos leer el .xlsx (¿archivo corrupto?)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            return Response({"detail": "El archivo está vacío."}, status=status.HTTP_400_BAD_REQUEST)

        # Mapa nombre_columna -> índice, tolerante a orden / mayúsculas / alias.
        col_index = {}
        for i, name in enumerate(header):
            key = _norm_header(name)
            if key:
                col_index[BULK_ALIASES.get(key, key)] = i
        missing = [c for c in BULK_REQUIRED_COLUMNS if c not in col_index]
        if missing:
            wb.close()
            return Response(
                {"detail": f"Faltan columnas obligatorias: {', '.join(missing)}."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        def cell(row, name):
            idx = col_index.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        # Lookups para validar FKs sin una consulta por fila.
        valid_categorias = set(Categoria.objects.values_list("id", flat=True))
        valid_cursos = set(Curso.objects.values_list("id", flat=True))
        valid_lecciones = set(Leccion.objects.values_list("id", flat=True))

        # Preguntas ya en BD (normalizadas) para saltar duplicados. La dedupe es
        # por texto de pregunta, igual que la limpieza manual del catálogo.
        existing_preguntas = set()
        if skip_duplicates:
            existing_preguntas = {
                (p or "").strip().lower()
                for p in Ejercicio.objects.values_list("pregunta", flat=True)
            }

        results = []
        to_create = []
        skipped = 0
        seen_preguntas = set()

        for offset, row in enumerate(rows):
            excel_row = offset + 2  # +1 cabecera, +1 base-1
            if row is None or all(_clean(v) is None for v in row):
                continue  # fila totalmente vacía

            pregunta = _clean(cell(row, "pregunta"))
            opciones = {k: _clean(cell(row, f"opcion_{k}")) for k in ("a", "b", "c", "d", "e", "f")}
            errors = []

            if not pregunta:
                errors.append("pregunta vacía")
            if not opciones["a"]:
                errors.append("opcion_a obligatoria")
            if not opciones["b"]:
                errors.append("opcion_b obligatoria")

            # Categoría (obligatoria).
            categoria_id = _as_int(cell(row, "categoria_id"))
            if categoria_id is None:
                errors.append("categoria_id obligatoria")
            elif categoria_id not in valid_categorias:
                errors.append(f"categoria_id {categoria_id} no existe")

            # Curso / lección (opcionales pero validados si vienen).
            curso_id = _as_int(cell(row, "curso_id"))
            if curso_id is not None and curso_id not in valid_cursos:
                errors.append(f"curso_id {curso_id} no existe")
                curso_id = None
            leccion_id = _as_int(cell(row, "leccion_id"))
            if leccion_id is not None and leccion_id not in valid_lecciones:
                errors.append(f"leccion_id {leccion_id} no existe")
                leccion_id = None

            # Respuesta: letra A-F o texto exacto de una opción.
            respuesta_text = None
            raw_resp = _clean(cell(row, "respuesta"))
            if not raw_resp:
                errors.append("respuesta vacía")
            else:
                letter = str(raw_resp).strip().lower()
                if letter in opciones and opciones[letter]:
                    respuesta_text = opciones[letter]
                else:
                    match = next(
                        (v for v in opciones.values() if v and v.lower() == str(raw_resp).lower()),
                        None,
                    )
                    if match:
                        respuesta_text = match
                    else:
                        errors.append(f"respuesta '{raw_resp}' no coincide con ninguna opción")

            if errors:
                results.append({"row": excel_row, "status": "error", "pregunta": (pregunta or "")[:80], "errors": errors})
                continue

            # Duplicado: mismo texto de pregunta ya visto en el archivo, o —si se
            # pidió— ya presente en BD.
            key = pregunta.lower() if pregunta else None
            in_file_dup = key is not None and key in seen_preguntas
            db_dup = skip_duplicates and key is not None and key in existing_preguntas
            dup = in_file_dup or db_dup
            if key:
                seen_preguntas.add(key)

            if skip_duplicates and dup:
                skipped += 1
                results.append({
                    "row": excel_row,
                    "status": "skipped",
                    "pregunta": (pregunta or "")[:80],
                    "errors": ["duplicado en BD" if db_dup else "duplicado en el archivo"],
                })
                continue

            imagen = _clean(cell(row, "imagen")) or "http://placeholder.url"
            obj = Ejercicio(
                categoria_id=categoria_id,
                curso_id=curso_id,
                leccion_id=leccion_id,
                pregunta=pregunta,
                imagen=imagen,
                opcion_a=opciones["a"], opcion_b=opciones["b"], opcion_c=opciones["c"],
                opcion_d=opciones["d"], opcion_e=opciones["e"], opcion_f=opciones["f"],
                respuesta=respuesta_text,
            )
            to_create.append((excel_row, pregunta, dup, obj))

        wb.close()

        created = 0
        if not dry_run and to_create:
            with db_transaction.atomic():
                Ejercicio.objects.bulk_create([o for _, _, _, o in to_create])
            created = len(to_create)

        for excel_row, pregunta, dup, obj in to_create:
            results.append({
                "row": excel_row,
                "status": "duplicate" if dup else ("ready" if dry_run else "created"),
                "pregunta": pregunta[:80],
                "id": obj.pk if not dry_run else None,
            })
        results.sort(key=lambda r: r["row"])

        error_count = sum(1 for r in results if r["status"] == "error")
        return Response({
            "dry_run": dry_run,
            "skip_duplicates": skip_duplicates,
            "total_rows": len(results),
            "created": created,
            "would_create": len(to_create) if dry_run else 0,
            "skipped": skipped,
            "errors": error_count,
            "results": results,
        }, status=status.HTTP_200_OK)


class GlosarioViewSet(viewsets.ModelViewSet):
    queryset = Glosario.objects.all()
    serializer_class = GlosarioSerializer
    permission_classes = [ReadOnlyOrAdmin]


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.all()
    serializer_class = CategoriaSerializer
    permission_classes = [PublicReadOrAdmin]


class RecursoViewSet(viewsets.ModelViewSet):
    """Biblioteca: GET /api/v1/schools/library/?categoria=&curso=&tipo=&q=

    - Admin/director: ven todo (director ve catálogo global; restricción de
      lectura por escuela no aplica al catálogo).
    - Estudiante: ve recursos donde `requires_owned_course=False` o donde el
      `curso` está en sus inscripciones (EstudianteCurso).
    """
    queryset = Recurso.objects.select_related('categoria', 'curso', 'leccion')
    serializer_class = RecursoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['categoria', 'curso', 'leccion', 'tipo', 'requires_owned_course']
    permission_classes = [ReadOnlyOrAdmin]

    def get_queryset(self):
        qs = super().get_queryset()
        # Búsqueda por título.
        q = self.request.query_params.get('q')
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(titulo__icontains=q) | Q(descripcion__icontains=q))

        user = self.request.user
        if is_admin(user) or is_director(user):
            return qs

        # Estudiante: ocultar recursos con requires_owned_course=True
        # excepto los de cursos que posee.
        if is_estudiante(user):
            from sales.models import EstudianteCurso
            owned_cursos = EstudianteCurso.objects.filter(
                estudiante_id=user
            ).values_list('curso_id_id', flat=True)
            from django.db.models import Q
            return qs.filter(
                Q(requires_owned_course=False) | Q(curso_id__in=owned_cursos)
            )

        return qs.none()


# Topes de subida (defensa del disco temporal ante llamadas directas a la API;
# el frontend valida antes). El temario es un índice pequeño; el material fuente
# suele ser un manual/libro completo y regularmente supera 25 MB.
TEMARIO_MAX_BYTES = 30 * 1024 * 1024       # 30 MB
CONTENIDO_MAX_BYTES = 300 * 1024 * 1024    # 300 MB


def _save_temp_pdf(uploaded) -> str:
    """Vuelca un archivo subido a un PDF temporal en disco y devuelve su ruta.

    PyMuPDF necesita una ruta de archivo, y el generador streamea durante un
    rato, así que persistimos la subida en vez de mantenerla en memoria.
    """
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    try:
        for chunk in uploaded.chunks():
            handle.write(chunk)
    finally:
        handle.close()
    return handle.name


class CourseGenerateView(APIView):
    """POST /api/v1/schools/courses/generate/

    Genera un curso completo (curso + unidades + lecciones) a partir de dos
    PDFs —TEMARIO (estructura) y CONTENIDO (fuente)— usando la app
    `content_pipeline`. Responde un stream NDJSON (`application/x-ndjson`) con
    eventos step/lesson/done/error consumido por `CourseGenerator.js`.

    Solo admin: crear cursos es una operación de back-office.
    """

    permission_classes = [IsAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        temario = request.FILES.get("temario")
        contenido = request.FILES.get("contenido")
        nombre = (request.data.get("nombre") or "").strip()
        codigo = (request.data.get("codigo") or "").strip().upper()

        if not temario or not contenido:
            return Response(
                {"detail": "Se requieren los PDFs 'temario' y 'contenido'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        oversized = None
        if temario.size > TEMARIO_MAX_BYTES:
            oversized = f"El temario supera el máximo de {TEMARIO_MAX_BYTES // (1024 * 1024)} MB."
        elif contenido.size > CONTENIDO_MAX_BYTES:
            oversized = f"El contenido supera el máximo de {CONTENIDO_MAX_BYTES // (1024 * 1024)} MB."
        if oversized:
            return Response({"detail": oversized}, status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE)
        if not nombre or not codigo:
            return Response(
                {"detail": "Se requieren 'nombre' y 'codigo'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if len(codigo) > 10:
            return Response(
                {"detail": "El código no puede superar 10 caracteres."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        is_profesional = str(request.data.get("is_profesional", "")).lower() in (
            "true", "1", "on", "yes",
        )
        try:
            max_lecciones = int(request.data.get("max_lecciones") or 20)
        except (TypeError, ValueError):
            max_lecciones = 20
        max_lecciones = max(1, min(max_lecciones, 100))
        idioma = (request.data.get("idioma") or "es").strip() or "es"
        modo = (request.data.get("modo") or "draft").strip() or "draft"
        source_name = getattr(contenido, "name", None) or f"Contenido: {nombre}"

        temario_path = _save_temp_pdf(temario)
        contenido_path = _save_temp_pdf(contenido)

        # Import diferido: mantiene los procesadores del pipeline fuera de la
        # ruta de importación del módulo de vistas.
        from content_pipeline.services.course_generator import generate_course_stream

        def stream():
            try:
                for event in generate_course_stream(
                    temario_path=temario_path,
                    contenido_path=contenido_path,
                    nombre=nombre,
                    codigo=codigo,
                    is_profesional=is_profesional,
                    max_lecciones=max_lecciones,
                    idioma=idioma,
                    modo=modo,
                    source_name=source_name,
                ):
                    yield json.dumps(event, ensure_ascii=False) + "\n"
            finally:
                for path in (temario_path, contenido_path):
                    try:
                        os.unlink(path)
                    except OSError:
                        pass

        response = StreamingHttpResponse(stream(), content_type="application/x-ndjson")
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"  # evita buffering en nginx
        return response

